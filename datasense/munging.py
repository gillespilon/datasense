"""
Data munging
"""

from typing import Callable, Dict, List, Optional, Tuple, Union, Pattern
from shutil import copytree, move, rmtree
from tkinter import filedialog
from warnings import warn
from pathlib import Path
from tkinter import Tk
from glob import glob
import textwrap
import string
import os

from datasense import random_data, timedelta_data, datetime_data
from openpyxl.worksheet.worksheet import Worksheet
from pandas.api.types import CategoricalDtype
from beautifultable import BeautifulTable
from scipy.stats import norm
import pandas as pd
import numpy as np


def dataframe_info(
    df: pd.DataFrame,
    file_in: Union[Path, str]
) -> pd.DataFrame:
    """
    Describe a dataframe.

    Display count of rows (rows_in_count)
    Display count of empty rows (rows_empty_count)
    Display count of non-empty rows (rows_out_count)
    Display count of columns (columns_in_count)
    Display count of empty columns (columns_empty_count)
    Display count of non-empty columns (columns_non_empty_count)
    Display table of data type, empty cell count, and empty cell percentage
        for non-empty columns
    Display count and list of non-empty columns
        (columns_non_empty_count, columns_non_empty_list)
    Display count and list of boolean columns
        (columns_bool_count, columns_bool_list)
    Display count and list of float columns
        (columns_float_count, columns_float_list)
    Display count and list of integer columns
        (columns_integer_count, columns_integer_list)
    Display count and list of datetime columns
        (columns_datetime_count, columns_datetime_list)
    Display count and list of string columns
        (columns_object_count, columns_object_list)
    Display count and list of timedelta columns
        (columns_timedelta_count, columns_timedelta_list)
    Display count and list of empty columns
        (columns_empty_count, columns_empty_list)

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    file_in : Union[Path, str]
        The name of the file from which df was created.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Example
    -------
    >>> df = ds.create_dataframe()
    >>> df = ds.dataframe_info(
    >>>     df=df,
    >>>     file_in='df'
    >>> )
    """
    df, rows_in_count, rows_out_count, rows_empty_count = process_rows(df)
    df, columns_in_count, columns_non_empty_count, columns_empty_count,\
        columns_empty_list, columns_non_empty_list, columns_bool_list,\
        columns_bool_count,\
        columns_float_list, columns_float_count,\
        columns_integer_list, columns_integer_count, columns_datetime_list,\
        columns_datetime_count, columns_object_list, columns_object_count,\
        columns_category_list, columns_category_count,\
        columns_timedelta_list, columns_timedelta_count\
        = process_columns(df=df)
    print('--------------------------')
    print(f'DataFrame information for: {file_in}')
    print()
    print(f'Rows total        : {rows_in_count}')
    print(f'Rows empty        : {rows_empty_count} (deleted)')
    print(f'Rows not empty    : {rows_out_count}')
    print(f'Columns total     : {columns_in_count}')
    print(f'Columns empty     : {columns_empty_count} (deleted)')
    print(f'Columns not empty : {columns_non_empty_count}')
    print()
    number_empty_cells_in_columns(df=df)
    print(f'List of {columns_non_empty_count} non-empty columns:')
    print_list_by_item(list=columns_non_empty_list)
    print()
    print(f'List of {columns_bool_count} bool columns:')
    print_list_by_item(list=columns_bool_list)
    print()
    print(f'List of {columns_category_count} category columns:')
    print_list_by_item(list=columns_category_list)
    print()
    print(f'List of {columns_datetime_count} datetime columns:')
    print_list_by_item(list=columns_datetime_list)
    print()
    print(f'List of {columns_float_count} float columns:')
    print_list_by_item(list=columns_float_list)
    print()
    print(f'List of {columns_integer_count} integer columns:')
    print_list_by_item(list=columns_integer_list)
    print()
    print(f'List of {columns_object_count} string columns:')
    print_list_by_item(list=columns_object_list)
    print()
    print(f'List of {columns_timedelta_count} timedelta columns:')
    print_list_by_item(list=columns_timedelta_list)
    print()
    print(f'List of {columns_empty_count} empty columns:')
    print_list_by_item(list=columns_empty_list)
    print()
    return df


def find_bool_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all boolean columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_bool : List[str]
        A list of boolean column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_bool = ds.find_bool_columns(df=df)
    >>> print(columns_bool)
    ['b']
    """
    columns_bool = list(df.select_dtypes(include=['bool', 'boolean']).columns)
    return columns_bool


def find_category_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all category columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_category : List[str]
        A list of category column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_category = ds.find_category_columns(df=df)
    >>> print(columns_category)
    ['c']
    """
    columns_category = list(df.select_dtypes(include=['category']).columns)
    return columns_category


def find_datetime_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all datetime columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_datetime : List[str]
        A list of datetime column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_datetime = ds.find_datetime_columns(df=df)
    >>> print(columns_datetime)
    ['t', 'u']
    """
    columns_datetime = list(df.select_dtypes(include=['datetime64']).columns)
    return columns_datetime


def find_float_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all float columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_float : List[str]
        A list of float column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_float = ds.find_float_columns(df=df)
    >>> print(columns_float)
    ['a', 'i', 'x', 'z']
    """
    columns_float = list(df.select_dtypes(include=['float64']).columns)
    return columns_float


def find_int_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all integer columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_int : List[str]
        A list of integer column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_int = ds.find_int_columns(df=df)
    >>> print(columns_int)
    ['y']
    """
    columns_int = list(df.select_dtypes(include=['int64', 'Int64']).columns)
    return columns_int


def find_int_float_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all integer and float columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_int_float : List[str]
        A list of integer and float column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_int_float = ds.find_int_float_columns(df=df)
    >>> print(columns_int_float)
    ['a', 'i', 'x', 'y', 'z']
    """
    columns_int_float = list(
        df.select_dtypes(include=['int64', 'float64']).columns
    )
    return columns_int_float


def find_object_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all object columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_object : List[str]
        A list of object column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_object = ds.find_object_columns(df=df)
    >>> print(columns_object)
    ['r', 's']
    """
    columns_object = list(df.select_dtypes(include=['object']).columns)
    return columns_object


def find_timedelta_columns(df: pd.DataFrame) -> List[str]:
    """
    Find all timedelta columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    columns_timedelta : List[str]
        A list of timedelta column names.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> columns_timedelta = ds.find_timedelta_columns(df=df)
    >>> print(columns_timedelta)
    ['d']
    """
    columns_timedelta = list(df.select_dtypes(include=['timedelta']).columns)
    return columns_timedelta


def number_empty_cells_in_columns(df: pd.DataFrame) -> None:
    """
    Create a table of data type, empty-cell count, and empty-all percentage
    for non-empty columns of a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Example
    -------
    >>> import datasense as ds
    >>> import pandas as pd
    >>> import numpy as np
    >>> df = pd.DataFrame({
    >>>     'X': [25.0, 24.0, 35.5, np.nan, 23.1],
    >>>     'Y': [27, 24, np.nan, 23, np.nan],
    >>>     'Z': ['a', 'b', np.nan, 'd', 'e']
    >>> })
    >>> empty_cells = ds.number_empty_cells_in_columns(df=df)
    >>> print(empty_cells)
    Information about non-empty columns
     Column   Data type   Empty cell count   Empty cell %   Unique
    -------- ----------- ------------------ -------------- --------
     X        float64                    1           20.0        4
     Y        float64                    2           40.0        3
     Z        object                     1           20.0        4
    """
    print('Information about non-empty columns')
    table = BeautifulTable(maxwidth=90)
    table.set_style(BeautifulTable.STYLE_COMPACT)
    column_alignments = {
        'Column': BeautifulTable.ALIGN_LEFT,
        'Data type': BeautifulTable.ALIGN_LEFT,
        'Empty cell count': BeautifulTable.ALIGN_RIGHT,
        'Empty cell %': BeautifulTable.ALIGN_RIGHT,
        'Unique': BeautifulTable.ALIGN_RIGHT,
    }
    table.columns.header = list(column_alignments.keys())
    for item, (_column_name, alignment) in\
            enumerate(column_alignments.items()):
        table.columns.alignment[item] = alignment
    num_rows = df.shape[0]
    for column_name in df:
        try:
            sum_nan = sum(pd.isnull(df[column_name]))
            percent_nan = round(sum_nan / num_rows * 100, 1)
            table.rows.append(
                [column_name,
                 df[column_name].dtype,
                 sum_nan,
                 percent_nan,
                 df[column_name].nunique()]
            )
        except KeyError:
            print('Error on column:', column_name)
    print(table)
    print()


def process_columns(df: pd.DataFrame) -> Tuple[
    pd.DataFrame,
    int,
    int,
    int,
    List[str],
    List[str],
    List[str],
    int,
    List[str],
    int,
    List[str],
    int,
    List[str],
    int,
    List[str],
    int,
    List[str],
    int,
    List[str],
    int,
]:
    """
    Create various counts of columns of a dataframe.

    Create count of columns
        (columns_in_count)
    Create count and list of empty columns
        (columns_empty_count, columns_empty_list)
    Create count and list of non-empty columns
        (columns_non_empty_count, columns_non_empty_list)
    Delete empty columns
    Create count and list of boolean columns
        (columns_bool_count, columns_bool_list)
    Create count and list of category columns
        (columns_category_count, columns_category_list)
    Create count and list of datetime columns
        (columns_datetime_count, columns_datetime_list)
    Create count and list of float columns
        (columns_float_count, columns_float_list)
    Create count and list of integer columns
        (columns_integer_count, columns_integer_list)
    Create count and list of string columns
        (columns_object_count, columns_object_list)
    Create count of timedelta columns
        (columns_timedelta_count, columns_timedelta_list)

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.
    columns_in_count : int
        The count of columns.
    columns_non_empty_count : int
        The count of non-empty columns.
    columns_empty_count: int
        The count of empty columns.
    columns_empty_list : List[str]
        The list of empty columns.
    columns_non_empty_list : List[str]
        The list of non-empty columns.
    columns_bool_list : List[str]
        The list of boolean columns.
    columns_bool_count : int
        The count of boolean columns.
    columns_float_list : List[str]
        The list of float columns.
    columns_float_count : int
        The count of float columns.
    columns_integer_list : List[str]
        The list of integer columns.
    columns_integer_count : int
        The count of integer columns
    columns_datetime_list : List[str]
        The list of datetime columns.
    columns_datetime_count : int
        The count of datetime columns.
    columns_object_list : List[str]
        The list of object columns.
    columns_object_count : int
        The count of object columns.
    columns_category_list : List[str]
        The list of category columns.
    columns_category_count : int
        The count of category columns.
    columns_timedelta_list : List[str]
        The list of timedelta columns.
    columns_timedelta_count : int
        The count of timedelta columns.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> df, columns_in_count, columns_non_empty_count, columns_empty_count,\
    >>>     columns_empty_list, columns_non_empty_list, columns_bool_list,\
    >>>     columns_bool_count, columns_float_list, columns_float_count,\
    >>>     columns_integer_list, columns_integer_count,\
    >>>     columns_datetime_list, columns_datetime_count,\
    >>>     columns_object_list, columns_object_count, columns_category_list,\
    >>>     columns_category_count, columns_timedelta_list,\
    >>>     columns_timedelta_count = ds.process_columns(df=df)
    columns_in_count       : 12
    columns_non_empty_count: 12
    columns_empty_count    : 0
    columns_empty_list     : []
    columns_non_empty_list :
        ['a', 'b', 'c', 'd', 'i', 'r', 's', 't', 'u', 'x', 'y', 'z']
    columns_bool_list      : ['b']
    columns_bool_count     : 1
    columns_float_list     : ['a', 'i', 'x', 'z']
    columns_float_count    : 4
    columns_integer_list   : ['y']
    columns_integer_count  : 1
    columns_datetime_list  : ['t', 'u']
    columns_datetime_count : 2
    columns_object_list    : ['r', 's']
    columns_object_count   : 2
    columns_category_list  : ['c']
    columns_category_count : 1
    columns_timedelta_list : ['d']
    columns_timedelta_count: 1
    """
    columns_empty_list = sorted({
        column_name for column_name in df.columns
        if df[column_name].isnull().all()
    })
    columns_in_count = len(df.columns)
    columns_empty_count = len(columns_empty_list)
    columns_non_empty_count = columns_in_count - columns_empty_count
    df = df.drop(columns_empty_list, axis='columns')
    columns_non_empty_list = sorted(df.columns)
    columns_bool_list = find_bool_columns(df=df)
    columns_bool_count = len(columns_bool_list)
    columns_category_list = find_category_columns(df=df)
    columns_category_count = len(columns_category_list)
    columns_datetime_list = find_datetime_columns(df=df)
    columns_datetime_count = len(columns_datetime_list)
    columns_float_list = find_float_columns(df=df)
    columns_float_count = len(columns_float_list)
    columns_integer_list = find_int_columns(df=df)
    columns_integer_count = len(columns_integer_list)
    columns_object_list = find_object_columns(df=df)
    columns_object_count = len(columns_object_list)
    columns_timedelta_list = find_timedelta_columns(df=df)
    columns_timedelta_count = len(columns_timedelta_list)
    return (
        df, columns_in_count, columns_non_empty_count,
        columns_empty_count, columns_empty_list, columns_non_empty_list,
        columns_bool_list, columns_bool_count,
        columns_float_list, columns_float_count,
        columns_integer_list, columns_integer_count,
        columns_datetime_list, columns_datetime_count,
        columns_object_list, columns_object_count,
        columns_category_list, columns_category_count,
        columns_timedelta_list, columns_timedelta_count
    )


def process_rows(df: pd.DataFrame) -> Tuple[pd.DataFrame, int, int, int]:
    """
    Create various counts of rows of a dataframe.
    Drop duplicate rows.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.
    rows_in_count : int
        The count of rows of the input dataframe.
    rows_out_count : int
        The count of rows of the output dataframe.
    rows_empty_count : int
        The count of empty rows of the input dataframe.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> df, rows_in_count, rows_out_count, rows_empty_count =\
    >>>     ds.process_rows(df)
    rows_in_count   : 42
    rows_out_count  : 42
    rows_empty_count: 0
    """
    rows_in_count = df.shape[0]
    df = df.dropna(axis='rows', how='all').drop_duplicates()
    rows_out_count = df.shape[0]
    rows_empty_count = rows_in_count - rows_out_count
    return (df, rows_in_count, rows_out_count, rows_empty_count)


def save_file(
    df: Union[pd.DataFrame, pd.Series],
    file_name: Union[str, Path],
    *,
    index: Optional[bool] = False,
    index_label: Optional[str] = None,
    sheet_name: Optional[str] = 'sheet_001',
) -> None:
    """
    Save a DataFrame or Series to a file.

    Parameters
    ----------
    df : Union[pd.DataFrame, pd.Series]
        The dataframe or series to be saved to a file.
    file_name : str
        The name of the file to be saved.
    index : Optional[bool]
        If True, creates an index.
    index_label : Optional[str]
        The index label.
    sheet_name : Optional[str]
        The name of the worksheet in the workbook.

    Examples
    --------
    Example 1
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> ds.save_file(
    >>>     df=df,
    >>>     file_name='x_y.csv'
    >>> )

    Example 2
    >>> ds.save_file(
    >>>     df=df,
    >>>     file_name='x_y.csv',
    >>>     index=True
    >>> )

    Example 3
    >>> ds.save_file(
    >>>     df=df,
    >>>     file_name='x_y.xlsx'
    >>> )

    Example 4
    >>> ds.save_file(
    >>>     df=df,
    >>>     file_name='x_y.xlsx',
    >>>     index=True,
    >>>     sheet_name='sheet_one'
    >>> )
    """
    if type(file_name).__name__ == 'str':
        file_name = Path(file_name)
    if file_name.suffix in ['.csv', '.CSV']:
        df.to_csv(
            path_or_buf=file_name,
            index=index,
            index_label=index_label
        )
    elif file_name.suffix in ['.ods', '.ODS']:
        excel_writer = pd.ExcelWriter(file_name)
        df.to_excel(
            excel_writer=excel_writer,
            sheet_name=sheet_name,
            engine='odf',
            index=index,
            index_label=index_label
        )
        excel_writer.save()
    elif file_name.suffix in ['.xlsx', '.XLSX']:
        excel_writer = pd.ExcelWriter(file_name)
        df.to_excel(
            excel_writer=excel_writer,
            sheet_name=sheet_name,
            engine='openpyxl',
            index=index,
            index_label=index_label
        )
        excel_writer.save()


def read_file(
    file_name: Union[str, Path],
    *,
    skiprows: Optional[List[int]] = None,
    column_names_dict: Optional[Dict[str, str]] = {},
    index_columns: Optional[List[str]] = [],
    usecols: Optional[List[str]] = None,
    dtype: Optional[dict] = None,
    converters: Optional[dict] = None,
    parse_dates: Optional[List[str]] = False,
    date_parser: Optional[Callable] = None,
    datetime_format: Optional[str] = None,
    time_delta_columns: Optional[List[str]] = [],
    category_columns: Optional[List[str]] = [],
    integer_columns: Optional[List[str]] = [],
    float_columns: Optional[List[str]] = [],
    boolean_columns: Optional[List[str]] = [],
    object_columns: Optional[List[str]] = [],
    sort_columns: Optional[List[str]] = [],
    sort_columns_bool: Optional[List[bool]] = [],
    sheet_name: Optional[str] = False,
    nrows: Optional[int] = None
) -> pd.DataFrame:
    """
    Create a dataframe from an external file.

    Parameters
    ----------
    file_name : str
        The name of the file to read.
    skiprows : Optional[List[int]]
        The specific row indices to skip.
    column_names_dict : Optional[List[str]]
        The new column names to replace the old column names.
    index_columns : Optional[List[str]]
        The columns to use for the dataframe index.
    usecols : Optional[List[str]] = None,
        The columns to read.
    dtype: Optional[dict] = None
        A dictionary of column names and dtypes.
    converters : Optional[dict] = None,
        Dictionary of functions for converting values in certain columns.
    parse_dates : Optional[List[str]] = False,
        The columns to use to parse date and time.
    date_parser : Optional[Callable] = None,
        The function to use for parsing date and time, when pandas needs
        extra help.
    datetime_format : Optional[str] = None,
        The str to use for formatting date and time.
    time_delta_columns : Optional[List[str]] = [],
        The columns to change to dtype timedelta.
    category_columns : Optional[List[str]] = []
        The columns to change to dtype category.
    integer_columns : Optional[List[str]] = []
        The columns to change to dtype integer.
    float_columns : Optional[List[str]] = []
        The columns to change to dtype float.
    boolean_columns : Optional[List[str]] = []
        The columns to change to dtype boolean.
    object_columns : Optional[List[str]] = []
        The columns to change to dtype object.
    sort_columns : Optional[List[str]] = []
        The columns on which to sort the dataframe.
    sort_columns_bool : Optional[List[bool]] = []
        The booleans for sort_columns.
    sheet_name : Optional[str] = False
        The name of the worksheet in the workbook.
    nrows : Optional[int] = None
        The number of rows to read.

    Returns
    -------
    df : pd.DataFrame
        The dataframe created from the external file.

    Examples
    --------
    Create a data file for the examples.
    >>> import datsense as ds
    >>> file_name='myfile.csv'
    >>> df = ds.create_dataframe()
    >>> print(df.columns)
    >>> print(df.dtypes)
    >>> ds.save_file(
    >>>     df=df,
    >>>     file_name=file_name
    >>> )
    Index(
        ['a', 'b', 'c', 'd', 'i', 'r', 's', 't', 'u', 'x', 'y', 'z'],
        dtype='object'
    )
    a            float64
    b            boolean
    c           category
    d    timedelta64[ns]
    i            float64
    r             object
    s             object
    t     datetime64[ns]
    u     datetime64[ns]
    x            float64
    y              Int64
    z            float64
    dtype: object

    # Example 1
    # Read a csv file. There is no guarantee the column dtypes will be correct.
    # Only [a, i, s, x, z] have the correct dtypes.
    >>> df = ds.read_file(file_name=file_name)
    >>> print(df.dtypes)
    a    float64
    b       bool
    c     object
    d     object
    i    float64
    r      int64
    s     object
    t     object
    u     object
    x    float64
    y      int64
    z    float64
    dtype: object

    # Example 2
    # Read a csv file. Ensure the dtypes of datetime columns.
    >>> parse_dates = ['t', 'u']
    >>> df = ds.read_file(
    >>>     file_name=file_name,
    >>>     parse_dates=parse_dates
    >>> )
    >>> print(df.dtypes)
    a           float64
    b              bool
    c            object
    d            object
    i           float64
    r             int64
    s            object
    t    datetime64[ns]
    u    datetime64[ns]
    x           float64
    y             int64
    z           float64
    dtype: object

    # Example 3
    # Read a csv file. Ensure the dtypes of columns; not timedelta, datetime.
    >>> convert_dict = {
    >>>     'a': 'float64',
    >>>     'b': 'boolean',
    >>>     'c': 'category',
    >>>     'i': 'float64',
    >>>     'r': 'str',
    >>>     's': 'str',
    >>>     'x': 'float64',
    >>>     'y': 'Int64',
    >>>     'z': 'float64'
    >>> }
    >>> df = ds.read_file(
    >>>     file_name=file_name,
    >>>     dtype=convert_dict
    >>> )
    >>> print(df.dtypes)
    a     float64
    b     boolean
    c    category
    d      object
    i     float64
    r      object
    s      object
    t      object
    u      object
    x     float64
    y       Int64
    z     float64
    dtype: object

    Example 4
    Read a csv file. Ensure the dtypes of columns. Rename the columns.
    Set index with another column. Convert float column to integer.
    >>> file_name = 'myfile.csv'
    >>> df = ds.create_dataframe()
    >>> ds.save_file(
    >>>     df=df,
    >>>     file_name=file_name
    >>> )
    >>> column_names_dict = {
    >>>     'a': 'A',
    >>>     'b': 'B',
    >>>     'c': 'C',
    >>>     'd': 'D',
    >>>     'i': 'I',
    >>>     'r': 'R',
    >>>     's': 'S',
    >>>     't': 'T',
    >>>     'u': 'U',
    >>>     'y': 'Y',
    >>>     'x': 'X',
    >>>     'z': 'Z'
    >>> }
    >>> index_columns = ['Y']
    >>> parse_dates = ['t', 'u']
    >>> time_delta_columns = ['D']
    >>> category_columns = ['C']
    >>> integer_columns = ['A', 'I']
    >>> float_columns = ['X']
    >>> boolean_columns = ['R']
    >>> object_columns = ['Z']
    >>> df = ds.read_file(
    >>>     file_name='myfile.csv',
    >>>     column_names_dict=column_names_dict,
    >>>     index_columns=index_columns,
    >>>     date_parser=date_parser(),
    >>>     parse_dates=parse_dates,
    >>>     time_delta_columns=time_delta_columns,
    >>>     category_columns=category_columns,
    >>>     integer_columns=integer_columns
    >>> )
    >>>
    >>>
    >>> def date_parser() -> Callable:
    >>>     return lambda s: datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
    >>>
    >>>
    >>> data = ds.read_file(
    >>>     file_name=file_name,
    >>>     column_names_dict=column_names_dict,
    >>>     index_columns=index_columns,
    >>>     date_time_columns=date_time_columns,
    >>>     date_parser=date_parser,
    >>>     parse_dates=date_time_columns,
    >>>     time_delta_columns=time_delta_columns,
    >>>     category_columns=category_columns,
    >>>     integer_columns=integer_columns
    >>> )

    Example 5
    Read an ods file.
    >>> file_name = 'myfile.ods'
    >>> df = ds.create_dataframe()
    >>> ds.save_file(
    >>>     df=df,
    >>>     file_name=file_name
    >>> )
    >>> parse_dates = ['t', 'u']
    >>> df = ds.read_file(
    >>>     file_name=file_name,
    >>>     parse_dates=parse_dates
    >>> )
    >>> ds.dataframe_info(
    >>>     df=df,
    >>>     file_in=file_name
    >>> )

    Example 6
    >>> Read an xlsx file.
    >>> file_name = 'mfile.xlsx'
    >>> sheet_name = 'raw_data'
    >>> df = ds.read_file(
    >>>     file_name=file_name,
    >>>     sheet_name=sheet_name
    >>> )
    >>> ds.dataframe_info(
    >>>     df=df,
    >>>     file_in=file_name
    >>> )
    """
    if type(file_name).__name__ == 'str':
        file_name = Path(file_name)
    if file_name.suffix in ['.csv', '.CSV']:
        df = pd.read_csv(
            file_name,
            skiprows=skiprows,
            usecols=usecols,
            dtype=dtype,
            converters=converters,
            parse_dates=parse_dates,
            date_parser=date_parser,
            nrows=nrows
        )
    elif file_name.suffix in ['.ods', '.ODS']:
        df = pd.read_excel(
            io=file_name,
            skiprows=skiprows,
            usecols=usecols,
            dtype=dtype,
            engine='odf',
            sheet_name=sheet_name,
            parse_dates=parse_dates,
            date_parser=date_parser
        )
    elif file_name.suffix in ['.xlsx', '.XLSX', '.xlsm', '.XLSM']:
        df = pd.read_excel(
            io=file_name,
            skiprows=skiprows,
            usecols=usecols,
            dtype=dtype,
            engine='openpyxl',
            sheet_name=sheet_name,
            parse_dates=parse_dates,
            date_parser=date_parser,
            nrows=nrows
        )
    if column_names_dict:
        df = rename_some_columns(
            df=df,
            column_names_dict=column_names_dict
        )
    if index_columns:
        df = df.set_index(index_columns)
    for column in category_columns:
        df[column] = df[column].astype(dtype=CategoricalDtype())
    for column in time_delta_columns:
        df[column] = pd.to_timedelta(df[column])
    for column in integer_columns:
        df[column] = df[column].astype(dtype='int64')
    for column in float_columns:
        df[column] = df[column].astype(dtype='float64')
    for column in boolean_columns:
        df[column] = df[column].astype(dtype='bool')
    for column in object_columns:
        df[column] = df[column].astype(dtype='object')
    if sort_columns and sort_columns_bool:
        df = sort_rows(
            df=df,
            sort_columns=sort_columns,
            sort_columns_bool=sort_columns_bool,
            kind='mergesort'
        )
    return df


def byte_size(
    num: np.int64,
    suffix: str = 'B'
) -> str:
    """
    Convert bytes to requested units.

    Parameters
    ----------
    num : np.int64
    suffix : str = 'B'

    Returns
    -------
    memory_usage : str

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.create_dataframe()
    >>> print(
    >>>     ds.byte_size(
    >>>         num=df.memory_usage(index=True).sum()
    >>>     )
    >>> )
    3.6 KiB
    """
    for unit in ['', 'Ki', 'Mi', 'Gi', 'Ti', 'Pi', 'Ei', 'Zi']:
        if abs(num) < 1024.0:
            return "%3.1f %s%s" % (num, unit, suffix)
        num /= 1024.0
    memory_usage = "%.1f %s%s" % (num, 'Yi', suffix)
    return memory_usage


def feature_percent_empty(
    df: pd.DataFrame,
    columns: List[str],
    threshold: float
) -> List[str]:
    """
    Remove features that have NaN > threshold.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    columns : List[str]
        The list of columns to evaluate.
    threshold : float
        The percentage empty threshold value.

    Returns
    -------
    list_columns : List[str]
        The list of columns below the threshold value.

    Example
    -------
    >>> import datasense as ds

    >>> features = ds.feature_percent_empty(
    >>>     df=data,
    >>>     columns=features,
    >>>     threshold=percent_empty_features
    >>> )
    """
    num_rows = df.shape[0]
    list_columns = [col for col in columns if
                    ((df[col].isna().sum() / num_rows * 100) <= threshold)]
    return list_columns


def create_directory(directories: List[str]) -> None:
    """
    Create empty directories for a path.

    Parameters
    ----------
    directories : List[str]
        The list of directories.

    Example
    -------
    >>> import datasense as ds
    >>> directory_list = ['directory_one', 'directory_two']
    >>> ds.create_directory(directories=directory_list)
    """
    for directory in directories:
        try:
            rmtree(directory)
        except Exception:
            pass
        Path(directory).mkdir(parents=True, exist_ok=True)


def delete_directory(directories: List[str]) -> None:
    """
    Delete a list of directories.

    Parameters
    ----------
    directories : List[str]
        The list of directories.

    Example
    -------
    >>> import datasense as ds
    >>> directory_list = ['directory_one', 'directory_two']
    >>> ds.delete_directory(directories=directory_list)
    """
    for directory in directories:
        try:
            rmtree(directory)
        except Exception:
            pass


def rename_directory(
    sources: List[str],
    destinations: List[str]
) -> None:
    """
    Delete destination directories (if present) and rename source directories
    to the destination directories.
    destination directory.

    Parameters
    ----------
    sources : List[str]
        The old directories.
    destinations : List[str]
        The new directories.

    Example
    -------
    >>> import datasense as ds
    >>> sources = ['old_directory']
    >>> destinations = ['new_directory']
    >>> ds.rename_directory(
    >>>     sources=sources,
    >>>     destinations=destinations
    >>> )
    """
    for source, destination in zip(sources, destinations):
        try:
            rmtree(destination)
        except Exception:
            pass
        move(
            src=source,
            dst=destination
        )


def copy_directory(
    sources: List[str],
    destinations: List[str]
) -> None:
    """
    Delte destination directories (if present) and copy source directories
    to destination directories.

    Parameters
    ----------
    sources : str
        The source directory name.
    destinations : str
        The destination directory name.

    Example
    -------
    >>> import datasense as ds
    >>> sources = ['source_directory']
    >>> destinations = ['destination_directory']
    >>> ds.rename_directory(
    >>>     sources=sources,
    >>>     destinations=destinations
    >>> )
    """
    for source, destination in zip(sources, destinations):
        try:
            rmtree(destination)
        except Exception:
            pass
        copytree(
            src=source,
            dst=destination
        )


def replace_text_numbers(
    df: pd.DataFrame,
    columns: Union[List[str], List[int], List[float], List[Pattern[str]]],
    old: Union[List[str], List[int], List[float], List[Pattern[str]]],
    new: List[int],
    *,
    regex: Optional[bool] = True
) -> pd.DataFrame:
    """
    Replace text or numbers with text or numbers.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    columns: Union[List[str], List[int], List[float], List[Pattern[str]]]
        The list of columns for replacement.
    old: Union[List[str], List[int], List[float], List[Pattern[str]]]
        The list of item to replace.
    new : List[int]
        The list of replacement items.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Examples
    -------
    Example 1
    >>> import datasense as ds
    >>> list_y_1_n_5 = [
    >>>     'Q01', 'Q02', 'Q03', 'Q04', 'Q05', 'Q06', 'Q10', 'Q17', 'Q18',
    >>>     'Q19', 'Q20', 'Q21', 'Q23', 'Q24', 'Q25'
    >>> ]
    >>> list_y_5_n_1 = [
    >>>     'Q07', 'Q11', 'Q12', 'Q13', 'Q15', 'Q16'
    >>> ]
    >>> data = ds.replace_text_numbers(
    >>>     df=data,
    >>>     columns=list_y_1_n_5,
    >>>     old=['Yes', 'No'],
    >>>     new=[1, 5],
    >>>     regex=False
    >>> )
    >>> data = ds.replace_text_numbers(
    >>>     df=data,
    >>>     columns=list_y_5_n_1,
    >>>     old=['Yes', 'No'],
    >>>     new=[5, 1],
    >>>     regex=False
    >>> )

    Example 2
    >>> data = ds.replace_text_numbers(
    >>>     df=data,
    >>>     columns=['Q23'],
    >>>     old=[r'\xa0'],
    >>>     new=[r' '],
    >>>     regex=True
    >>> )

    Example 3
    >>> data = ds.replace_text_numbers(
    >>>     df=data,
    >>>     columns=['address_country'],
    >>>     old=[
    >>>         'AD', 'AE', 'AF', 'AG',
    >>>         'AI', 'AL', 'AM', 'AN',
    >>>         'AO', 'AQ', 'AR', 'AS',
    >>>         'AT', 'AU', 'AW', 'AZ',
    >>>     ]
    >>>     new=[
    >>>         'Andorra', 'Unit.Arab Emir.', 'Afghanistan', 'Antigua/Barbuda',
    >>>         'Anguilla', 'Albania', 'Armenia', 'Niederl.Antill.',
    >>>         'Angola', 'Antarctica', 'Argentina', 'Samoa,American',
    >>>         'Austria', 'Australia', 'Aruba', 'Azerbaijan',
    >>>     ],
    >>>     regex=False
    >>> )
    """
    dfnew = df.copy(deep=True)
    for column in columns:
        dfnew[column] = dfnew[column].replace(
            to_replace=old,
            value=new,
            regex=regex
        )
    return dfnew


def create_dataframe(
    *,
    size: Optional[int] = 42
) -> pd.DataFrame:
    # TODO: why did I create distribution "u"?
    """
    Create a Pandas dataframe.

    Parameters
    ----------
    size : Optional[int] = 42
        The number of rows to create.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Notes
    -----
    a  : float64
    b  : bool
    bn : boolean (nullable)
    c  : category
    cs : CategoricalDtype category
    d  : timedelta64[ns]
    r  : object
    s  : object
    t  : datetime64[ns]
    u  : datetime64[ns]
    x  : float64
    y  : int64
    yn : Int64
    z  : float64

    Example
    -------
    >>> df = create_dataframe()
    """
    df = pd.DataFrame(
        {
            'a': random_data(
                distribution='uniform',
                size=size,
                loc=13,
                scale=70
            ),
            'b': random_data(
                distribution='bool',
                size=size
            ),
            'bn': random_data(
                distribution='boolean',
                size=size
            ),
            'c': random_data(
                distribution='category',
                size=size
            ),
            'cs': random_data(
                distribution='categories',
                size=size
            ),
            'd': timedelta_data(time_delta_days=size-1),
            'r': random_data(
                distribution='strings',
                strings=['0', '1'],
                size=size
            ),
            's': random_data(
                distribution='strings',
                size=size
                ),
            't': datetime_data(time_delta_days=size-1),
            'u': datetime_data(time_delta_days=size-1),
            'x': random_data(
                distribution='norm',
                size=size
            ),
            'y': random_data(
                distribution='randint',
                size=size
            ),
            'yn': random_data(
                distribution='randInt',
                size=size
            ),
            'z': random_data(
                distribution='uniform',
                size=size
            )
        }
    )
    return df


def create_dataframe_norm(
    *,
    row_count: Optional[int] = 42,
    column_count: Optional[int] = 13,
    loc: Optional[float] = 69,
    scale: Optional[float] = 13,
    random_state: Optional[int] = None,
    column_names: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    Create dataframe of random normal data.

    Parameters
    ----------
    row_count : Optional[int] = 42,
        The number of rows to create.
    column_count : Optional[int] = 13,
        The number of columns to create.
    loc : Optional[float] = 69,
        The mean of the data.
    scale : Optional[float] = 13
        The standard deviation of the data.
    random_state: Optional[int] = None
        The random number seed.
    column_names: Optional[List[str]]
        The column names.

    Examples
    --------
    Example 1
    >>> df = ds.create_dataframe_norm()

    Example 2
    >>> column_names = [f'col{item}' for item in range(column_count)]
    >>> row_count = 1000
    >>> column_count = 100
    >>> df = ds.create_dataframe_norm(
    >>>     row_count=row_count,
    >>>     column_count=column_count,
    >>>     loc=69,
    >>>     scale=13,
    >>>     random_state=42,
    >>>     column_names=column_names
    >>> )
    """
    if not column_names:
        column_names = [f'col{item}' for item in range(column_count)]
    df = pd.DataFrame(
        norm.rvs(
            size=(row_count, column_count),
            loc=loc,
            scale=scale,
            random_state=random_state
        ),
        columns=column_names
    )
    return df


def delete_rows(
    df: pd.DataFrame,
    delete_row_criteria:
        Union[Tuple[str, int], Tuple[str, float], Tuple[str, str]]
) -> pd.DataFrame:
    """
    Delete rows of a dataframe based on a value in one column.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    delete_row_criteria : Tuple[str, int]
        A tuple of column name and criteria for the entire cell.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.delete_rows(
    >>>     df=df,
    >>>     delete_row_criteria=['Batch Acceptance', 1]
    >>> )
    """

    if delete_row_criteria:
        df = df.loc[~(df[delete_row_criteria[0]] == delete_row_criteria[1])]
    return df


def delete_columns(
    df: pd.DataFrame,
    columns: List[str]
) -> pd.DataFrame:
    """
    Delete columns of a dataframe using a list.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    columns : List[str]
        A list of column names.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.delete_columns(
    >>>     df=df,
    >>>     columns=columns
    >>> )
    """
    df = df.drop(columns=columns)
    return df


def sort_rows(
    df: pd.DataFrame,
    sort_columns: List[str],
    sort_columns_bool: List[bool],
    kind: str = 'mergesort'
) -> pd.DataFrame:
    """
    Sort a dataframe for one or more columns.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    sort_columns : List[str]
        The sort columns.
    sort_columns_bool : List[bool]
        The booleans for sort_columns: True = ascending, False = descending.
    kind: str = 'mergesort'
        The sort algorithm.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.sort_rows(
    >>>     df=df,
    >>>     sort_columns=sort_columns,
    >>>     sort_columns_bool=sort_columns_bool,
    >>>     kind='mergesort'
    >>> )
    """

    df = df.sort_values(
        by=sort_columns,
        axis='index',
        ascending=sort_columns_bool,
        kind=kind
    )
    return df


def rename_all_columns(
    df: pd.DataFrame,
    labels: List[str]
) -> pd.DataFrame:
    """
    Rename all dataframe columns.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    labels : List[str]
        The list of all column names.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.rename_all_columns(
    >>>     df=df,
    >>>     labels=labels
    >>> )
    """
    df = df.set_axis(
        labels=labels,
        axis='columns'
    )
    return df


def rename_some_columns(
    df: pd.DataFrame,
    column_names_dict: Dict[str, str]
) -> pd.DataFrame:
    """
    Rename some columns with a dictionary.

    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe.
    column_names_dict : Dict[str, str]
        The dictionary of old:new column names.

    Returns
    -------
    df : pd.DataFrame
        The output dataframe.

    Example
    -------
    >>> import datasense as ds
    >>> df = ds.rename_some_columns(
    >>>     df=df,
    >>>     column_names_dict=column_names_dict
    >>> )
    """
    df = df.rename(columns=column_names_dict)
    return df


def replace_column_values(
    s: pd.Series,
    replace_dict: Dict[str, str]
) -> pd.Series:
    """
    Replace values in a series using a dictionary.

    Parameters
    ----------
    s : pd.Series
        The input series.
    replace_dict : Union[Dict[str, str], Dict[int, int], Dict[float, float]]
        The dictionary of values to replace.

    Returns:
    s : pd.Series
        The output series.

    Examples
    -------=
    >>> import datasense as ds
    >>> df = ds.replace_column_values(
    >>>     df=df,
    >>>     column=column,
    >>>     replace_dict=replace_dict
    >>> )
    """
    s = s.replace(
        to_replace=replace_dict,
        regex=True,
        value=None
    )
    return s


def directory_file_list(
    directory: Union[str, Path],
    *,
    patterns: Optional[List[str]] = None
) -> List[Path]:
    """
    Return a list of files within a directory.

    Parameters
    ----------
    path : Union[str, Path]
        The path of the directory.
    extension : List[str]
        The file extensions to use for finding files in the path.

    Returns
    -------
    files : List[Path]
        A list of paths.

    Examples
    --------
    Example 1
    ---------
    >>> patterns = ['.pdf']
    >>> import datasense as ds
    >>> files = ds.directory_file_list(
    >>>     directory=path,
    >>>     patterns=patterns
    >>> )

    Example 2
    ---------
    >>> patterns = ['.PDF']
    >>> import datasense as ds
    >>> files = ds.directory_file_list(
    >>>     directory=path,
    >>>     patterns=patterns
    >>> )

    Example 3
    ---------
    >>> patterns = ['.pdf', '.PDF']
    >>> import datasense as ds
    >>> files = ds.directory_file_list(
    >>>     directory=path,
    >>>     patterns=patterns
    >>> )

    Example 4
    ---------
    >>> # Return all files within a directory
    >>> import datasense as ds
    >>> files = ds.directory_file_list(
    >>>     directory=path
    >>> )
    """
    directory = Path(directory)
    if not patterns:
        files = [
            x for x in directory.iterdir()
        ]
    else:
        files = [
            x for x in directory.iterdir() if x.suffix in patterns
        ]
    return files


def directory_file_print(
    directory: Union[str, Path],
    *,
    text: Optional[str] = 'Files in directory'
) -> None:
    """
    Print the files in a path.

    Parameters
    ----------
    path : Path
        The path of the files to print.
    text : Optional[str]
        The text to print.

    Example
    -------
    >>> import datasense as ds
    >>> path = <path to a directory>
    >>> text = 'your text'
    >>> ds.directory_file_print(
    >>>     path=path,
    >>>     text=text
    >>> )
    """
    directory = Path(directory)
    if text:
        print(f'{text}', directory)
    for x in directory.iterdir():
        print(x.name)
    print()


def directory_remove_file(
    path: Path,
    file_names: List[str]
) -> List[str]:
    """
    Parameters
    ----------
    path : Path
        The path of the file to remove.
    file_names : List[str]
        The list of files from which to remove the path.

    Returns
    -------
    file_names : {list[str]
        The list of files without the removed path.

    Example
    -------
    >>> import datasense as ds
    >>> file_names = ds.directory_remove_file(
    >>>     path=path,
    >>>     file_names=file_names
    >>> )
    """
    for file in file_names:
        if path.name in file:
            os.remove(file)
            file_names.remove(file)
    return file_names


def print_list_by_item(
    list: List[str],
    *,
    title: Optional[str] = None,
    width: Optional[int] = 80
) -> None:
    '''
    Print each item of a list.

    Parameters
    ----------
    list : List[str]
        The list of strings to print.
    width : Optional[int] = 80
        The width of the output in characters.

    Example
    -------
    >>> import datasense as ds
    >>> ds.print_list_by_item(list=my_list_to_print)
    '''
    wrapper = textwrap.TextWrapper(width=width)
    string_not_list = ", ".join(list)
    new_list = wrapper.wrap(string_not_list)
    if title:
        print(title)
    for element in new_list:
        print(element)


def ask_directory_path(
    *,
    title: Optional[str] = 'Select directory',
    initialdir: Optional[Path] = None,
    print_bool: Optional[bool] = False
) -> Path:
    """
    Ask user for directory.

    Parameters
    ----------
    title : Optional[str] = 'Select directory'
        The title of the dialog window.
    initialdir : Optional[Path]
        The directory in which the dialogue starts.
    print_bool : Optional[bool] = True
        A boolean. Print message if True.

    Returns
    -------
    path: Path
        The path of the directory.

    Example
    -------
    >>> from tkinter import filedialog
    >>> from pathlib import Path
    >>> from tkinter import Tk
    >>> import datasense as ds
    >>> path = ds.ask_directory_path(title='your message')
    """
    rootwindow = Tk()
    path = filedialog.askdirectory(
        parent=rootwindow,
        initialdir=initialdir,
        title=title
    )
    path = Path(path)
    rootwindow.destroy()
    if print_bool:
        print(title)
        print(path)
        print()
    return path


def ask_open_file_name_path(
    title: str,
    *,
    initialdir: Optional[Path] = None,
    filetypes: Optional[List[Tuple[str]]] = [('xlsx files', '.xlsx .XLSX')]
) -> Path:
    """
    Ask user for the path of the file to open.

    Parameters
    ----------
    title : str
        The title of the dialog window.
    initialdir : Optional[Path]
        The directory in which the dialogue starts.
    filetypes : Optional[List[Tuple[str]]] = [('xlsx files', '.xlsx .XLSX')]
        The file types to make visible.

    Returns
    -------
    path: Path
        The path of the file to open.

    Examples
    --------
    Example 1
    >>> from tkinter import filedialog
    >>> from pathlib import Path
    >>> from tkinter import Tk
    >>> import datasense as ds
    >>> path = ds.ask_open_file_name_path(title='your message')

    Example 2
    >>> path = ds.ask_open_file_name_path(
    >>>     title='your message',
    >>>     filetypes=[('csv files', '.csv .CSV')]
    >>> )
    """
    rootwindow = Tk()
    path = filedialog.askopenfilename(
        parent=rootwindow,
        title=title,
        initialdir=initialdir,
        filetypes=filetypes
    )
    path = Path(path)
    rootwindow.destroy()
    return path


def ask_save_as_file_name_path(
    title: str,
    *,
    initialdir: Optional[Path] = None,
    filetypes: Optional[List[Tuple[str]]] = [('xlsx files', '.xlsx .XLSX')],
    print_bool: Optional[bool] = True
) -> Path:
    """
    Ask user for the path of the file to save as.

    Parameters
    ----------
    title : Optional[str] = 'Select file'
        The title of the dialog window.
    initialdir : Optional[Path]
        The directory in which the dialogue starts.
    print_bool : Optional[bool] = True
        A boolean. Print message if True.

    Returns
    -------
    path: Path
        The path of the file to save as.

    Example2
    --------
    Example 1
    >>> from tkinter import filedialog
    >>> from pathlib import Path
    >>> from tkinter import Tk
    >>> import datasense as ds
    >>> path = ds.ask_save_as_file_name_path(title='your message')

    Example 2
    >>> path = ds.ask_save_as_file_name_path(
    >>>     title='your message',
    >>>     filetypes=[('csv files', '.csv .CSV')]
    >>> )
    """
    rootwindow = Tk()
    path = filedialog.asksaveasfilename(
        parent=rootwindow,
        title=title,
        filetypes=filetypes
    )
    path = Path(path)
    rootwindow.destroy()
    if print_bool:
        print(title)
        print(path)
        print()
    return path


def series_replace_string(
    series: pd.Series,
    find: str,
    replace: str,
    *,
    regex: Optional[bool] = True
) -> pd.Series:
    """
    Find and replace a string in a series.

    Parameters
    ----------
    series : pd.Series
        The input series of data.
    find : str
        The string to find.
    replace : str
        The replacement string.
    regex : Optional[bool] = True
        Determines if the passed-in pattern is a regular expression.

    Returns
    -------
    series : pd.Series
        The output series of data.

    Example
    -------
    >>> import datasense as ds
    >>> df[column] = series_replace_string(
    >>>     series=df[column],
    >>>     find='find this text',
    >>>     replace='replace with this text'
    >>> )
    """
    series = series.str.replace(
        pat=find,
        repl=replace,
        regex=regex
    )
    return series


def unique_list_items(
    ws: Worksheet,
    row_of_labels: int,
    row_below_labels: int,
    column_name_varname: str,
    text_to_replace: List[str],
    text_to_remove: List[str]
) -> Tuple[List[int], List[int]]:
    '''
    Determine list of unique items in varname.
    Replace text.

    TODO:
    This function does too many things. Break it up.
    Add detail to docstring.
    '''
    column_numbers = [
        col for col in range(ws.min_column, ws.max_column + 1)
    ]
    column_labels = []
    for row in ws.iter_rows(
        min_row=row_of_labels,
        max_row=row_of_labels,
        min_col=column_numbers[0],
        max_col=column_numbers[-1]
    ):
        for cell in row:
            column_labels.append(cell.value)
    column_names_numbers = dict(zip(column_labels, column_numbers))
    varname_replace = []
    for row in ws.iter_rows(
        min_row=row_below_labels,
        min_col=column_names_numbers[column_name_varname],
        max_col=column_names_numbers[column_name_varname]
    ):
        for cell in row:
            if cell.value:
                varname_replace.append(cell.value)
    varname_remove = []
    for item in varname_replace:
        for thing in text_to_replace:
            item = str(item).replace(thing, '')
        varname_remove.append(item)
    varname = [x for x in varname_remove if x not in text_to_remove]
    varname = (list(set(varname)))
    return (varname, column_numbers)


def list_directories_within_directory(
    path: Union[str, Path]
) -> List[str]:
    '''
    Return a list of directories found within a path.

    Parameters
    ----------
    path : Union[str, Path]
        The path of the enclosing directory.

    Returns
    -------
    directory_list : List[str]
        A list of directories.

    Example
    -------
    >>> import datasense as ds
    >>> directory_list = ds.list_directories_within_directory(
    >>>     path='directory_companies'
    >>> )
    '''
    path = Path(path)
    directory_list = [item.name for item in path.iterdir() if item.is_dir()]
    return directory_list


def remove_punctuation(list_dirty: List[str]) -> List[str]:
    '''
    Remove punctuation from list items.

    Parameters
    ----------
    list_dirty : List[str]
        The list of items containing punctuation.

    Returns
    -------
    list_clean : List[str]
        The list of items without punctuation.

    Example
    -------
    >>> import datasense as ds
    >>> list_clean = ds.remove_punctuation(list_dirty=list_dirty)
    '''
    list_clean = [
        ''.join(
            character for character in item
            if character not in string.punctuation
        )
        for item in list_dirty
    ]
    return list_clean


def list_change_case(
    list_dirty: List[str],
    case: str
) -> List[str]:
    '''
    Change the case of items in a list.

    Parameters
    ----------
    list_dirty : List[str]
        The list of strings.
    case : str
        The type of case to apply.

    Returns
    -------
    list_clean : List[str]
        The list of strings with case applied.

    Example
    -------
    >>> import datasense as ds
    >>> list_clean = ds.list_change_case(
    >>>     list_dirty=list_dirty,
    >>>     case='upper'
    >>> )
    '''
    if case == 'upper':
        list_clean = [x.upper() for x in list_dirty]
    elif case == 'lower':
        list_clean = [x.lower() for x in list_dirty]
    elif case == 'title':
        list_clean = [x.title() for x in list_dirty]
    elif case == 'capitalize':
        list_clean = [x.capitalize() for x in list_dirty]
    return list_clean


def listone_contains_all_listtwo_substrings(
    listone: List[str],
    listtwo: List[str]
) -> List[str]:
    '''
    Return a list of items from one list that contain substrings of items
    from another list.

    Parameters
    ----------
    listone : List[str]
        The list of items in which there are substrings to match from listtwo.
    listwo : List[str]
        The list of items that are substrings of the items in listone.

    Returns
    -------
    matches : List[str]
        The list of items from listone that contain substrings of the items
        from listtwo.

    Example
    -------
    >>> listone = ['prefix-2020-21-CMJG-suffix', 'bobs your uncle']
    >>> listwo = [ 'CMJG', '2020-21']
    >>> matches = ds.listone_contains_all_listwo_substrings(
    >>>     listone=listone,
    >>>     listtwo=listtow
    >>> )
    >>> ['prefix-2020-21-CMJG-suffix']
    '''
    matches = [x for x in listone if all(y in x for y in listtwo)]
    return matches


def list_one_list_two_ops(
    list_one: Union[List[str], List[int], List[float]],
    list_two: Union[List[str], List[int], List[float]],
    action: str
) -> Union[List[str], List[int], List[float]]:
    '''
    Create list of items comparing two lists:
    - Items unique to list_one
    - Items unique to list_two
    - Items common to both lists

    Parameters
    ----------
    list_one : Union[List[str], List[int], List[float]]
        A list of items.
    list_two : Union[List[str], List[int], List[float]]
        A list of items.
    action : str
        A string of either "list_one", "list_two", or "intersection"

    Returns
    -------
    list_result : Union[List[str], List[int], List[float]]
        The list of unique items.

    Examples
    --------
    Example 1
    >>> import datasense as ds
    >>> list_one = [1, 2, 3, 4, 5, 6]
    >>> list_two = [4, 5, 6, 7, 8, 9]
    >>> list_one_unique = ds.list_one_unique_list_two(
    >>>     list_one=list_one,
    >>>     list_two=list_two,
    >>>     action='list_one'
    >>> )
    >>> [1, 2, 3]

    Example 2
    >>> list_one = [1, 2, 3, 4, 5, 6]
    >>> list_two = [4, 5, 6, 7, 8, 9]
    >>> list_one_unique = ds.list_one_unique_list_two(
    >>>     list_one=list_one,
    >>>     list_two=list_two,
    >>>     action='list_two'
    >>> )
    >>> [7, 8, 9]

    Example 3
    >>> list_one = [1, 2, 3, 4, 5, 6]
    >>> list_two = [4, 5, 6, 7, 8, 9]
    >>> list_one_unique = ds.list_one_unique_list_two(
    >>>     list_one=list_one,
    >>>     list_two=list_two,
    >>>     action='intersection'
    >>> )
    >>> [4, 5, 6]
    '''
    if action == 'list_one':
        list_result = list(set(list_one).difference(list_two))
    elif action == 'list_two':
        list_result = list(set(list_two).difference(list_one))
    elif action == 'intersection':
        list_result = list(set(list_one).intersection(list_two))
    else:
        print('Error. Enter "list_one" or "list_two" for parameter "unique"')
    return list_result


__all__ = (
    'dataframe_info',
    'find_bool_columns',
    'find_category_columns',
    'find_datetime_columns',
    'find_float_columns',
    'find_int_columns',
    'find_int_float_columns',
    'find_object_columns',
    'find_timedelta_columns',
    'number_empty_cells_in_columns',
    'process_columns',
    'process_rows',
    'read_file',
    'save_file',
    'byte_size',
    'feature_percent_empty',
    'create_directory',
    'delete_directory',
    'rename_directory',
    'copy_directory',
    'replace_text_numbers',
    'create_dataframe',
    'create_dataframe_norm',
    'delete_rows',
    'delete_columns',
    'sort_rows',
    'rename_all_columns',
    'rename_some_columns',
    'replace_column_values',
    'directory_file_list',
    'directory_file_print',
    'directory_remove_file',
    'print_list_by_item',
    'ask_directory_path',
    'ask_open_file_name_path',
    'ask_save_as_file_name_path',
    'series_replace_string',
    'unique_list_items',
    'list_directories_within_directory',
    'remove_punctuation',
    'list_change_case',
    'listone_contains_all_listtwo_substrings',
    'list_one_list_two_ops',
)
